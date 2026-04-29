import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# Create a new workbook and select the active worksheet
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Functional Test Cases"

# Define the headers typically found in a standard Functional Test Cases document
headers = [
    "Test Case ID",
    "Module",
    "Priority",
    "Test Description",
    "Pre-conditions",
    "Test Steps",
    "Expected Result",
    "Actual Result",
    "Status",
    "Comments"
]

# Write headers to the first row and apply styling
header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)

for col_num, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_num, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = alignment

# Define the test cases based on Graft's architecture (Auth, Dashboard, Editor, Deployment Flow, Extension Sync)
test_cases = [
    # Auth Module
    ("TC_AUTH_01", "Authentication", "High", "Verify successful GitHub OAuth login", "User has a valid GitHub account", "1. Navigate to landing page.\n2. Click 'Get Started for Free'.\n3. Authenticate with GitHub.", "User is redirected to the Dashboard successfully.", "Redirected to dashboard successfully", "Pass", ""),
    ("TC_AUTH_02", "Authentication", "Medium", "Verify behavior of 'Get Started' button for logged-in users", "User is already logged in", "1. Navigate to landing page.\n2. Verify button text.", "Button text should be 'Build now' and lead directly to the Dashboard.", "Button correctly formatted and redirects", "Pass", ""),
    
    # Dashboard / Team Management
    ("TC_DASH_01", "Dashboard", "High", "Verify creating a new Team", "User is logged in", "1. Go to Dashboard.\n2. Click 'Create Team'.\n3. Enter valid team name.\n4. Submit.", "New team is created and user is directed to team view.", "Team created successfully", "Pass", ""),
    ("TC_DASH_02", "Dashboard", "High", "Verify listing scripts for a specific team", "Team exists with scripts", "1. Open a specific team's dashboard.\n2. Observe the script list.", "All scripts belonging to the team are displayed with current versions and statuses.", "Scripts match DB status", "Pass", ""),
    
    # Editor / Script Creation
    ("TC_EDIT_01", "Editor", "High", "Verify creating a new script", "User is in a team dashboard", "1. Click 'Create Script'.\n2. Fill in Name, Target URLs, and initial Code.\n3. Save.", "Script is saved, version 1 is created, and user is taken to the Monaco Editor view.", "Script saves and loads successfully", "Pass", ""),
    ("TC_EDIT_02", "Editor", "Medium", "Verify Monaco Editor loads without ChunkLoadError", "Script exists", "1. Open an existing script in Editor context.", "Monaco Editor initializes correctly and displays the script's code.", "Editor loads flawlessly", "Pass", ""),
    ("TC_EDIT_03", "Editor", "Low", "Verify Analytics Tab initializes safely when empty", "Script has no analytics data", "1. Open Editor.\n2. Click Analytics tab.", "Analytics tab loads with 0s without throwing 'Cannot read properties of undefined' error.", "Page renders without error 500", "Pass", "Fixed via backend API patch"),
    
    # Deployment Flow (Test -> Review -> Live)
    ("TC_DEP_01", "Deployment", "High", "Verify 'Test new changes' button behavior", "Browser linked via Companion", "1. Make a code change in Editor.\n2. Click 'Test new changes'.", "Script status changes to 'testing'. Test version is deployed only to author's browser.", "Code saved and tested successfully", "Pass", "RLS bug bypassed successfully"),
    ("TC_DEP_02", "Deployment", "Medium", "Verify error when testing without a linked browser", "No Companion extension linked", "1. Click 'Test new changes'.", "Returns error: 'No browser linked for testing. Open your Companion popup...'", "Correct error shown", "Pass", ""),
    ("TC_DEP_03", "Deployment", "High", "Verify Submitting a script for review", "Script is in 'testing' state", "1. Click 'Submit for Review'.", "Script status transitions securely from 'testing' to 'pending_review' via adminDb bypass.", "Status goes to pending_review", "Pass", ""),
    ("TC_DEP_04", "Deployment", "High", "Verify Admin Approving a script", "Script is in 'pending_review'. User is Admin.", "1. Log in as Team Admin.\n2. Click 'Approve' on script.", "Script status transitions to 'live' and is prepared for team-wide Companion sync.", "Script turned Live successfully", "Pass", ""),
    ("TC_DEP_05", "Deployment", "High", "Verify Admin Rejecting a script", "Script is 'pending_review'. User is Admin.", "1. Log in as Team Admin.\n2. Click 'Reject' and provide reason.", "Script status transitions to 'rejected' and rejection reason is saved.", "Script rejected and saved reason", "Pass", ""),
    
    # Extension Sync
    ("TC_EXT_01", "Extension", "High", "Verify extension polling fetches 'live' scripts", "Extension installed, user linked, script is 'live'", "1. Wait 30 seconds for polling cycle.\n2. Inspect extension payload.", "Extension successfully retrieves the newly approved live script from the API endpoint.", "", "Unexecuted", ""),
    ("TC_EXT_02", "Extension", "High", "Verify script execution on matching Target URL", "Script synced to extension", "1. Navigate to URL matching target_urls array.\n2. Observe page behavior.", "Script executes successfully within the sandboxed extension context.", "", "Unexecuted", ""),
    ("TC_EXT_03", "Extension", "Medium", "Verify live configuration (remote_config) sync", "Script is live, remote_config updated", "1. Modify script config via API.\n2. Wait 30 seconds.", "Extension fetches new config without requiring a full code redeploy.", "", "Unexecuted", "")
]

# Write data to rows
for row_num, row_data in enumerate(test_cases, 2):
    for col_num, cell_value in enumerate(row_data, 1):
        cell = ws.cell(row=row_num, column=col_num, value=cell_value)
        cell.alignment = left_align
        # Center align ID, Module, Priority, Status
        if col_num in [1, 2, 3, 9]:
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Adjust column widths for readability
column_widths = {
    "A": 15,  # ID
    "B": 15,  # Module
    "C": 12,  # Priority
    "D": 35,  # Description
    "E": 25,  # Pre-conditions
    "F": 40,  # Steps
    "G": 35,  # Expected
    "H": 20,  # Actual
    "I": 15,  # Status
    "J": 20   # Comments
}

for col_letter, width in column_widths.items():
    ws.column_dimensions[col_letter].width = width

# Save the workbook
output_path = r"D:/Coding/Functional_Test_Cases_Graft.xlsx"
wb.save(output_path)
print(f"Successfully generated Test Cases Excel sheet at {output_path}")
